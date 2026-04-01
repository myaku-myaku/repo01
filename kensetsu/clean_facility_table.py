from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import numbers
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore


DATE_RE = re.compile(r"(\d{4}/\d{1,2}/\d{1,2})")


@dataclass(frozen=True)
class Row:
    date_str: str
    purpose: str


def build_purpose_to_dates(rows: Iterable[Row]) -> List[Tuple[str, List[datetime.date]]]:
    """Return purposes mapped to sorted unique dates (as date objects)."""

    purpose_to_dates: dict[str, set[datetime.date]] = {}
    for r in rows:
        purpose_to_dates.setdefault(r.purpose, set()).add(_parse_date(r.date_str))

    items: List[Tuple[str, List[datetime.date]]] = []
    for purpose, dates in purpose_to_dates.items():
        items.append((purpose, sorted(dates)))

    # Stable ordering: most frequent purposes first, then alphabetical
    items.sort(key=lambda x: (-len(x[1]), x[0]))
    return items


def _normalize_whitespace(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"\s+", " ", text).strip()


def extract_date_purpose_pairs(raw_text: str) -> List[Row]:
    """Extract sequential (date, purpose) pairs from arbitrary text.

    Rules:
    - Find every date like YYYY/M/D (or YYYY/MM/DD).
    - Purpose is the text after that date until the next date.
    - Collapse whitespace in purpose.
    """

    text = raw_text.replace("\ufeff", "")
    matches = list(DATE_RE.finditer(text))
    rows: List[Row] = []

    for index, match in enumerate(matches):
        date_str = match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        purpose = text[start:end]
        purpose = purpose.strip().strip('"')
        purpose = _normalize_whitespace(purpose)
        if purpose:
            rows.append(Row(date_str=date_str, purpose=purpose))

    return rows


def write_csv(rows: Iterable[Row], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["日付", "目的"])
        for r in rows:
            writer.writerow([r.date_str, r.purpose])


def _parse_date(date_str: str):
    return datetime.strptime(date_str, "%Y/%m/%d").date()


def write_xlsx(rows: Iterable[Row], out_path: Path) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not available in this environment")

    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows_list = list(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整理済"

    ws.append(["日付", "目的"])
    for r in rows_list:
        ws.append([_parse_date(r.date_str), r.purpose])

    # format
    for cell in ws["A"][1:]:
        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60

    # Pivot-like sheet: purposes as headers, dates listed below
    ws_pivot = wb.create_sheet(title="目的別")
    purpose_to_dates = build_purpose_to_dates(rows_list)

    for col_index, (purpose, dates) in enumerate(purpose_to_dates, start=1):
        header_cell = ws_pivot.cell(row=1, column=col_index, value=purpose)
        ws_pivot.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 28

        for row_index, date_value in enumerate(dates, start=2):
            cell = ws_pivot.cell(row=row_index, column=col_index, value=date_value)
            cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    wb.save(out_path)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Split facility schedule text into (date, purpose) rows")
    parser.add_argument("input", type=Path, help="Input file path (e.g. kensetsu/ファシリティ管理課.ini)")
    parser.add_argument("--csv", type=Path, default=None, help="Output CSV path")
    parser.add_argument("--xlsx", type=Path, default=None, help="Output XLSX path")
    args = parser.parse_args(argv)

    raw_text = args.input.read_text(encoding="utf-8", errors="replace")
    rows = extract_date_purpose_pairs(raw_text)

    base = args.input.with_suffix("")
    out_csv = args.csv or Path(f"{base}_cleaned.csv")
    out_xlsx = args.xlsx or Path(f"{base}_cleaned.xlsx")

    write_csv(rows, out_csv)

    # XLSX is optional; if openpyxl isn't available, skip with a clear message.
    try:
        write_xlsx(rows, out_xlsx)
    except Exception as e:
        print(f"⚠ xlsx出力をスキップしました: {e}")

    print(f"✅ 出力: {out_csv}")
    if out_xlsx.exists():
        print(f"✅ 出力: {out_xlsx}")
    print(f"件数: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
