#!/usr/bin/env python3
"""Scan Excel files for CAB hostnames and 準伝送ラック mentions."""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
import re
from pathlib import Path
from typing import Dict, Iterable, List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

JUNDENSO_KEYWORD = "準伝送ラック"
CAB_PATTERN = re.compile(r"\b\S*CAB\d+[A-Za-z]\S*\b", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Extract CAB hostnames and 準伝送ラック occurrences from Excel files.",
	)
	parser.add_argument(
		"--dir",
		type=Path,
		default=Path(__file__).resolve().parent,
		help="Directory to scan (recursively). Defaults to the script directory.",
	)
	parser.add_argument(
		"--output",
		type=Path,
		default=None,
		help="Optional path for the result workbook (.xlsx).",
	)
	return parser.parse_args()


def collect_excel_files(target_dir: Path) -> List[Path]:
	exts = (".xlsx", ".xlsm")
	files = [p for p in target_dir.rglob("*") if p.suffix.lower() in exts and p.is_file()]
	return sorted(files)


def scan_excel_file(file_path: Path) -> Dict[str, object]:
	cab_hostnames: Set[str] = set()
	jundenso_cells: List[Dict[str, str]] = []
	sheet_hits = defaultdict(int)
	errors: List[str] = []

	try:
		wb = load_workbook(filename=file_path, read_only=True, data_only=True)
	except (InvalidFileException, OSError) as exc:
		errors.append(str(exc))
		return {
			"file": file_path,
			"cab_hostnames": cab_hostnames,
			"jundenso_cells": jundenso_cells,
			"sheet_hits": sheet_hits,
			"errors": errors,
		}

	with wb:
		for sheet in wb.worksheets:
			try:
				for row in sheet.iter_rows(values_only=False):
					for cell in row:
						value = cell.value
						if not isinstance(value, str):
							continue
						normalized = value.strip()
						if not normalized:
							continue
						if JUNDENSO_KEYWORD in normalized:
							count = normalized.count(JUNDENSO_KEYWORD)
							sheet_hits[sheet.title] += count
							jundenso_cells.append(
								{
									"ファイル名": file_path.name,
									"シート名": sheet.title,
									"セル番地": cell.coordinate,
									"値": normalized,
								}
							)
						for match in CAB_PATTERN.findall(normalized):
							cab_hostnames.add(match)
			except Exception as exc:  # noqa: BLE001
				errors.append(f"{sheet.title}: {exc}")

	return {
		"file": file_path,
		"cab_hostnames": cab_hostnames,
		"jundenso_cells": jundenso_cells,
		"sheet_hits": sheet_hits,
		"errors": errors,
	}


def build_output_frames(results: Iterable[Dict[str, object]]) -> Dict[str, pd.DataFrame]:
	summary_rows = []
	cell_rows: List[Dict[str, str]] = []
	host_rows: List[Dict[str, str]] = []

	for item in results:
		file_path: Path = item["file"]
		cab_hostnames: Set[str] = item["cab_hostnames"]
		sheet_hits: Dict[str, int] = item["sheet_hits"]
		jundenso_cells = item["jundenso_cells"]
		errors = item["errors"]

		if sheet_hits:
			summary_rows.append(
				{
					"ファイル名": file_path.name,
					"準伝送シート数": len(sheet_hits),
					"準伝送セル件数": sum(sheet_hits.values()),
					"CABホスト数": len(cab_hostnames),
					"CABホスト名": ", ".join(sorted(cab_hostnames)),
					"エラー": " / ".join(errors),
				}
			)
		elif errors:
			summary_rows.append(
				{
					"ファイル名": file_path.name,
					"準伝送シート数": 0,
					"準伝送セル件数": 0,
					"CABホスト数": len(cab_hostnames),
					"CABホスト名": ", ".join(sorted(cab_hostnames)),
					"エラー": " / ".join(errors),
				}
			)

		cell_rows.extend(jundenso_cells)

		for host in sorted(cab_hostnames):
			rack_match = re.search(r"CAB(\d+)([A-Za-z])", host, re.IGNORECASE)
			host_rows.append(
				{
					"ファイル名": file_path.name,
					"ホスト名": host,
					"ラック番号": f"CAB{rack_match.group(1)}" if rack_match else "",
					"末尾": rack_match.group(2) if rack_match else "",
					"ネットワークラック": "Yes" if rack_match and rack_match.group(2).lower() == "n" else "No",
				}
			)

	return {
		"summary": pd.DataFrame(summary_rows),
		"cells": pd.DataFrame(cell_rows),
		"hosts": pd.DataFrame(host_rows),
	}


def main() -> None:
	args = parse_args()
	target_dir = args.dir.expanduser().resolve()
	if not target_dir.exists():
		print(f"Target directory not found: {target_dir}")
		return

	excel_files = collect_excel_files(target_dir)
	if not excel_files:
		print("No Excel files (.xlsx/.xlsm) were found.")
		return

	print(f"📁 対象フォルダ: {target_dir}")
	print(f"📊 検出されたExcelファイル: {len(excel_files)}件")

	results: List[Dict[str, object]] = []
	for idx, file_path in enumerate(excel_files, 1):
		print(f"[{idx}/{len(excel_files)}] {file_path}", flush=True)
		item = scan_excel_file(file_path)
		results.append(item)
		sheet_count = len(item["sheet_hits"])
		if sheet_count:
			print(
				f"  -> 準伝送ラック: {sheet_count}シート / {sum(item['sheet_hits'].values())}セル, CABホスト: {len(item['cab_hostnames'])}件",
				flush=True,
			)
		elif item["errors"]:
			print(f"  -> ⚠ 失敗: {' / '.join(item['errors'])}", flush=True)
		else:
			print("  -> 該当なし", flush=True)

	frames = build_output_frames(results)
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	output_path = (
		args.output
		if args.output is not None
		else Path(__file__).resolve().parent / f"準伝送ラック抽出_{timestamp}.xlsx"
	)

	with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
		if not frames["summary"].empty:
			frames["summary"].to_excel(writer, sheet_name="準伝送ファイル一覧", index=False)
		else:
			pd.DataFrame([{"メッセージ": "準伝送ラックが含まれるファイルはありませんでした"}]).to_excel(
				writer,
				sheet_name="準伝送ファイル一覧",
				index=False,
			)

		if not frames["cells"].empty:
			frames["cells"].to_excel(writer, sheet_name="準伝送セル詳細", index=False)
		if not frames["hosts"].empty:
			frames["hosts"].to_excel(writer, sheet_name="CABホスト名一覧", index=False)

	print(f"✅ 完了: {output_path}")


if __name__ == "__main__":
	main()