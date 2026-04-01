import pandas as pd
import openpyxl
from pathlib import Path
import re
from collections import defaultdict
from datetime import datetime

JUNDENSO_KEYWORD = '準伝送ラック'

def extract_cab_hostnames_and_jundenso(file_path):
    """
    Excelファイルから'CAB'を含むホスト名を抽出
    
    Args:
        file_path: Excelファイルのパス
    
    Returns:
        tuple[set, int]: (CABを含むホスト名のセット, 準伝送ラックの出現数)
    """
    cab_hostnames = set()
    jundenso_count = 0
    
    try:
        # openpyxlで読み込み（数式ではなく値を取得）
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value and isinstance(cell_value, str):
                        if JUNDENSO_KEYWORD in cell_value:
                            jundenso_count += 1
                        # 'CAB'を含む文字列を検出
                        if 'CAB' in cell_value.upper():
                            # ホスト名のパターンにマッチするか確認
                            # 例: FKOL163b-CAB02n
                            match = re.search(r'\S*CAB\d+[a-zA-Z]\S*', cell_value, re.IGNORECASE)
                            if match:
                                cab_hostnames.add(match.group(0))
        
        wb.close()
        
    except Exception as e:
        print(f"  ⚠ エラー: {e}")
    
    return cab_hostnames, jundenso_count


def analyze_cab_racks(hostnames):
    """
    CABホスト名を分析してラック情報を集計
    
    Args:
        hostnames: ホスト名のセット
    
    Returns:
        dict: ラック番号ごとの情報
    """
    rack_info = defaultdict(lambda: {'total': 0, 'n_suffix': 0, 'hostnames': set()})
    
    for hostname in hostnames:
        # CAB番号と末尾の文字を抽出
        # 例: FKOL163b-CAB02n -> CAB02, n
        match = re.search(r'CAB(\d+)([a-zA-Z])', hostname, re.IGNORECASE)
        if match:
            rack_num = match.group(1)  # "02"
            suffix = match.group(2).lower()  # "n"
            
            rack_info[rack_num]['hostnames'].add(hostname)
            rack_info[rack_num]['total'] = len(rack_info[rack_num]['hostnames'])
            
            if suffix == 'n':
                rack_info[rack_num]['n_suffix'] += 1
    
    return rack_info


def main():
    # 対象フォルダ（スクリプトと同じディレクトリ）
    target_folder = Path(__file__).parent
    
    # xlsxとxlsmファイルを検索
    excel_files = list(target_folder.glob('*.xlsx')) + list(target_folder.glob('*.xlsm'))
    
    if not excel_files:
        print("❌ Excelファイルが見つかりませんでした")
        return
    
    print(f"📁 対象フォルダ: {target_folder}")
    print(f"📊 検出されたExcelファイル: {len(excel_files)}件\n")
    
    # 全拠点のデータを収集
    all_data = []
    all_cab_hostnames = set()
    jundenso_file_count = 0
    jundenso_total_count = 0
    
    for idx, file_path in enumerate(excel_files, 1):
        print(f"[{idx}/{len(excel_files)}] 処理中: {file_path.name}")
        
        # CABホスト名を抽出 + 準伝送ラック検出
        cab_hostnames, jundenso_count = extract_cab_hostnames_and_jundenso(file_path)
        has_jundenso = jundenso_count > 0
        if has_jundenso:
            jundenso_file_count += 1
            jundenso_total_count += jundenso_count
        
        if cab_hostnames:
            # ラック情報を分析
            rack_info = analyze_cab_racks(cab_hostnames)
            
            total_racks = len(rack_info)
            total_n_suffix = sum(info['n_suffix'] for info in rack_info.values())
            
            all_data.append({
                '拠点ファイル名': file_path.stem,
                'ラック本数': total_racks,
                'ネットワークラック数(末尾n)': total_n_suffix,
                '準伝送ラックあり': 'Yes' if has_jundenso else 'No',
                '準伝送ラック数': jundenso_count,
                'CABホスト名': ', '.join(sorted(cab_hostnames))
            })
            
            all_cab_hostnames.update(cab_hostnames)
            
            extra = f", 準伝送ラック: {jundenso_count}件" if has_jundenso else ""
            print(f"  ✓ ラック: {total_racks}本, ネットワークラック: {total_n_suffix}本, ホスト名: {len(cab_hostnames)}件{extra}")
        else:
            extra = f"（準伝送ラック: {jundenso_count}件）" if has_jundenso else ""
            print(f"  - CABホスト名なし{extra}")

            # CABが無くても準伝送ラック情報は残す
            if has_jundenso:
                all_data.append({
                    '拠点ファイル名': file_path.stem,
                    'ラック本数': 0,
                    'ネットワークラック数(末尾n)': 0,
                    '準伝送ラックあり': 'Yes',
                    '準伝送ラック数': jundenso_count,
                    'CABホスト名': ''
                })
    
    # 結果をDataFrameに変換
    df_result = pd.DataFrame(all_data)
    
    # 全体サマリーを作成
    summary_data = {
        '項目': ['総ラック本数', '総ネットワークラック数', 'ユニークホスト名数', '拠点数', '準伝送ラックあり拠点数', '準伝送ラック総件数'],
        '値': [
            df_result['ラック本数'].sum() if not df_result.empty else 0,
            df_result['ネットワークラック数(末尾n)'].sum() if not df_result.empty else 0,
            len(all_cab_hostnames),
            len(excel_files),
            jundenso_file_count,
            jundenso_total_count,
        ]
    }
    df_summary = pd.DataFrame(summary_data)
    
    # ユニークホスト名リスト
    unique_hostnames_data = []
    for hostname in sorted(all_cab_hostnames):
        match = re.search(r'CAB(\d+)([a-zA-Z])', hostname, re.IGNORECASE)
        if match:
            rack_num = match.group(1)
            suffix = match.group(2)
            unique_hostnames_data.append({
                'ホスト名': hostname,
                'ラック番号': f'CAB{rack_num}',
                '末尾': suffix,
                'ネットワークラック': 'Yes' if suffix.lower() == 'n' else 'No'
            })
    
    df_hostnames = pd.DataFrame(unique_hostnames_data)
    
    # Excelファイルに出力
    jundenso_tag = '_準伝送ラックあり' if jundenso_file_count > 0 else ''
    output_file = target_folder / f'CABラック集計結果{jundenso_tag}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='サマリー', index=False)
        df_result.to_excel(writer, sheet_name='拠点別集計', index=False)
        df_hostnames.to_excel(writer, sheet_name='全ホスト名一覧', index=False)
    
    print(f"\n✅ 完了！結果を保存しました: {output_file.name}")
    print(f"\n📈 集計結果:")
    print(f"  - 総ラック本数: {df_result['ラック本数'].sum() if not df_result.empty else 0}本")
    print(f"  - ネットワークラック: {df_result['ネットワークラック数(末尾n)'].sum() if not df_result.empty else 0}本")
    print(f"  - ユニークホスト名: {len(all_cab_hostnames)}件")
    print(f"  - 準伝送ラックあり拠点数: {jundenso_file_count}件")
    print(f"  - 準伝送ラック総件数: {jundenso_total_count}件")
    print(f"  - 処理拠点数: {len(excel_files)}件")


if __name__ == '__main__':
    try:
        main()
    except BrokenPipeError:
        # Allow piping output (e.g., `| head`) without stack traces.
        import os
        import sys

        try:
            sys.stdout = open(os.devnull, 'w')
        except Exception:
            pass
