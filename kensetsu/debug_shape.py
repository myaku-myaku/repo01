#!/usr/bin/env python3
"""Debug script to inspect shape structure in D111_沼南.xlsx"""

from pathlib import Path
from openpyxl import load_workbook

JUNDENSO_KEYWORD = "準伝送ラック"

file_path = Path(__file__).parent / "D111_沼南.xlsx"

if not file_path.exists():
    print(f"ファイルが見つかりません: {file_path}")
    exit(1)

print(f"📁 検査中: {file_path}\n")

wb = load_workbook(filename=file_path, read_only=False, data_only=False)

for sheet in wb.worksheets:
    print(f"\n{'='*60}")
    print(f"シート: {sheet.title}")
    print(f"{'='*60}")
    
    # セル値をチェック
    found_in_cells = False
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and JUNDENSO_KEYWORD in value:
                found_in_cells = True
                print(f"  ✓ セル内で発見: {value[:50]}...")
    
    if not found_in_cells:
        print("  - セル内には該当なし")
    
    # すべての属性を調査
    print(f"\n  Sheet attributes: {[a for a in dir(sheet) if not a.startswith('__')]}")
    
    # _drawing属性を確認
    if hasattr(sheet, "_drawing") and sheet._drawing:
        drawing = sheet._drawing
        print(f"\n  _drawing: 存在")
        print(f"  _drawing type: {type(drawing)}")
        
        # すべての属性を列挙
        for attr in dir(drawing):
            if not attr.startswith('_'):
                val = getattr(drawing, attr, None)
                if val and not callable(val):
                    print(f"    {attr}: {type(val)} = {val if len(str(val)) < 100 else str(val)[:100]+'...'}")
    else:
        print("\n  _drawing: None または存在しない")
    
    # コメントをチェック
    if hasattr(sheet, "_comments") and sheet._comments:
        print(f"\n  コメント数: {len(sheet._comments)}")
        for comment in sheet._comments:
            if hasattr(comment, "text") and JUNDENSO_KEYWORD in str(comment.text):
                print(f"    ✓ コメントで発見: {comment.text}")
    
    # Legacy drawing (VML)
    if hasattr(sheet, "_legacy_drawing"):
        print(f"\n  _legacy_drawing: {sheet._legacy_drawing}")
    
    # OLE objects
    if hasattr(sheet, "_charts"):
        print(f"\n  _charts: {sheet._charts}")
    
    # Data validations
    if hasattr(sheet, "data_validations"):
        print(f"\n  data_validations: {sheet.data_validations}")
    
    # Tables
    if hasattr(sheet, "_tables"):
        print(f"\n  _tables: {sheet._tables}")
    
    # Images
    if hasattr(sheet, "_images"):
        print(f"\n  _images: {sheet._images}")

# Workbookレベルのチェック
print(f"\n{'='*60}")
print("Workbook レベル")
print(f"{'='*60}")

if hasattr(wb, "_drawings"):
    print(f"wb._drawings: {wb._drawings}")

if hasattr(wb, "vba_archive"):
    print(f"wb.vba_archive: {wb.vba_archive}")

# 全シートのdrawingをチェック
print(f"\nワークブック内の全描画オブジェクト:")
for sheet in wb.worksheets:
    if hasattr(sheet, "_drawing") and sheet._drawing is not None:
        print(f"  {sheet.title}: drawing exists")
        drawing = sheet._drawing
        
        # Anchorをチェック（図形の配置情報）
        if hasattr(drawing, "twoCellAnchor"):
            anchors = drawing.twoCellAnchor if isinstance(drawing.twoCellAnchor, list) else [drawing.twoCellAnchor]
            print(f"    twoCellAnchors: {len(anchors)}")
            for idx, anchor in enumerate(anchors):
                print(f"      Anchor {idx}: {type(anchor)}")
                for attr in dir(anchor):
                    if not attr.startswith('_'):
                        val = getattr(anchor, attr, None)
                        if val and not callable(val):
                            print(f"        {attr}: {val}")
    else:
        print(f"  {sheet.title}: no drawing")

wb.close()
print("\n\n✅ 検査完了")
