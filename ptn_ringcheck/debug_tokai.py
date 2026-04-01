#!/usr/bin/env python3
"""Debug script to check FSR9 sheet in PTN_Tunnel_PW(東海).xlsm"""

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook

file_path = Path("/Users/Taishi.Kurosaki/Documents/repo01/repo01/ptn_ringcheck/PTN_Tunnel_PW/PTN_Tunnel_PW(東海).xlsm")

NNI_PATTERN = re.compile(r"NNI\s*\d{1,3}G", re.IGNORECASE)
LAG_PATTERN = re.compile(r"\b\S*LAG\S*\b", re.IGNORECASE)

print("=== Drawing XML ===")
try:
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        drawing_files = [n for n in zip_ref.namelist() if 'xl/drawings/drawing' in n and n.endswith('.xml')]
        
        for drawing_file in drawing_files:
            content = zip_ref.read(drawing_file).decode('utf-8', errors='replace')
            root = ET.fromstring(content)
            for elem in root.iter():
                if elem.text and isinstance(elem.text, str):
                    text = elem.text.strip()
                    if text and ('NNI' in text.upper() or 'LAG' in text.upper()):
                        nni_matches = NNI_PATTERN.findall(text)
                        lag_matches = LAG_PATTERN.findall(text)
                        if nni_matches or lag_matches:
                            print(f"{drawing_file}: '{text}'")
                            print(f"  NNI: {nni_matches}, LAG: {lag_matches}")
except Exception as exc:
    print(f"Error: {exc}")

print("\n=== FSR9 Sheet ===")
wb = load_workbook(filename=file_path, read_only=True, data_only=True)
for sheet in wb.worksheets:
    if sheet.title == "FSR9":
        print(f"Found sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str) and ('NNI' in value.upper() or 'LAG' in value.upper()):
                    nni_matches = NNI_PATTERN.findall(value)
                    lag_matches = LAG_PATTERN.findall(value)
                    if nni_matches or lag_matches:
                        print(f"  Cell: '{value}'")
                        print(f"    NNI: {nni_matches}, LAG: {lag_matches}")
wb.close()
