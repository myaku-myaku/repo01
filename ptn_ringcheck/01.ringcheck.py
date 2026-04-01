#!/usr/bin/env python3

import csv
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Set
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


# NNI + 数字 + G のパターン (例: NNI100G, NNI20G)
NNI_PATTERN = re.compile(r"NNI\s*\d{1,3}G", re.IGNORECASE)
# LAGを含むパターン (例: downlink LAG, LAG#1)
LAG_PATTERN = re.compile(r"\b\S*LAG\S*\b", re.IGNORECASE)


def collect_excel_files(target_dir: Path) -> List[Path]:
    exts = (".xlsx", ".xlsm")
    files = [
        p for p in target_dir.rglob("*") 
        if p.suffix.lower() in exts 
        and p.is_file()
        and not p.name.startswith("~$")  # Excelの一時ファイルを除外
        and not p.name.startswith(".")   # 隠しファイルを除外
    ]
    return sorted(files)


def extract_sheet_data(file_path: Path) -> List[dict]:
    rows = []
    sheet_matches = {}  # sheet_name -> {"nni": set(), "lag": set()}
    drawing_to_sheet = {}  # drawing file -> sheet name のマッピング
    
    # Drawing XMLとシートの関係を解析
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # workbook.xmlからシート名とsheetIdの対応を取得
            wb_content = zip_ref.read('xl/workbook.xml').decode('utf-8', errors='replace')
            wb_root = ET.fromstring(wb_content)
            sheet_id_to_name = {}
            for sheet_elem in wb_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                sheet_name = sheet_elem.get('name')
                r_id = sheet_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                if sheet_name and r_id:
                    sheet_id_to_name[r_id] = sheet_name
            
            # xl/_rels/workbook.xml.rels からrIdとsheet*.xmlの対応を取得
            workbook_rels = zip_ref.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='replace')
            rels_root = ET.fromstring(workbook_rels)
            rid_to_sheet_file = {}
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                r_id = rel.get('Id')
                target = rel.get('Target')
                if r_id and target and 'worksheets/sheet' in target:
                    # worksheets/sheet1.xml -> sheet1.xml
                    sheet_file = target.split('/')[-1].replace('.xml', '')
                    rid_to_sheet_file[r_id] = sheet_file
            
            # rIdからシート名へのマッピングを構築
            rid_to_name = {}
            for r_id, sheet_name in sheet_id_to_name.items():
                if r_id in rid_to_sheet_file:
                    rid_to_name[rid_to_sheet_file[r_id]] = sheet_name
            
            # xl/worksheets/_rels/sheet*.xml.rels を解析してdrawingとシートの対応を取得
            rels_files = [n for n in zip_ref.namelist() if 'xl/worksheets/_rels/sheet' in n and n.endswith('.rels')]
            
            for rels_file in rels_files:
                try:
                    # xl/worksheets/_rels/sheet1.xml.rels -> sheet1
                    sheet_file = rels_file.split('/')[-1].replace('.xml.rels', '')
                    sheet_name = rid_to_name.get(sheet_file, f'{sheet_file} (Unknown)')
                    
                    content = zip_ref.read(rels_file).decode('utf-8', errors='replace')
                    root = ET.fromstring(content)
                    
                    for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        target = rel.get('Target')
                        if target and '../drawings/drawing' in target:
                            # ../drawings/drawing1.xml -> drawing1.xml
                            drawing_name = target.split('/')[-1]
                            drawing_to_sheet[f'xl/drawings/{drawing_name}'] = sheet_name
                except Exception:
                    pass
    except Exception:
        pass
    
    # ZIPとして開いてdrawing XMLを直接スキャン
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            drawing_files = [n for n in zip_ref.namelist() if 'xl/drawings/drawing' in n and n.endswith('.xml')]
            
            for drawing_file in drawing_files:
                try:
                    content = zip_ref.read(drawing_file).decode('utf-8', errors='replace')
                    root = ET.fromstring(content)
                    
                    # このDrawing XMLが属するシートを特定
                    sheet_name = drawing_to_sheet.get(drawing_file, f'{drawing_file} (Unknown)')
                    
                    for elem in root.iter():
                        if elem.text and isinstance(elem.text, str):
                            text = elem.text.strip()
                            if text:
                                # NNIパターンのマッチ
                                for match in NNI_PATTERN.findall(text):
                                    if sheet_name not in sheet_matches:
                                        sheet_matches[sheet_name] = {"nni": set(), "lag": set()}
                                    sheet_matches[sheet_name]["nni"].add(match)
                                # LAGパターンのマッチ
                                for match in LAG_PATTERN.findall(text):
                                    if sheet_name not in sheet_matches:
                                        sheet_matches[sheet_name] = {"nni": set(), "lag": set()}
                                    sheet_matches[sheet_name]["lag"].add(match)
                except Exception as exc:
                    print(f"    ! Error reading {drawing_file} in {file_path.name}: {exc}", flush=True)
    except Exception:
        pass  # ZIP読み込み失敗は無視してopenpyxlでの処理を続行

    # openpyxlでセル値をスキャン
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    except (InvalidFileException, OSError) as exc:
        print(f"  ! Failed to open {file_path.name}: {exc}", flush=True)
        # Drawing XMLで見つかったマッチがあれば出力
        for sheet_name, matches in sheet_matches.items():
            nni_list = "; ".join(sorted(matches["nni"]))
            lag_list = "; ".join(sorted(matches["lag"]))
            rows.append({
                "file": str(file_path),
                "sheet_name": sheet_name,
                "nni_matches": nni_list,
                "lag_matches": lag_list,
            })
        return rows

    try:
        for sheet in wb.worksheets:
            if sheet.title not in sheet_matches:
                sheet_matches[sheet.title] = {"nni": set(), "lag": set()}
            try:
                for row in sheet.iter_rows(values_only=True):
                    for value in row:
                        if isinstance(value, str):
                            # NNIパターンのマッチ
                            for match in NNI_PATTERN.findall(value):
                                sheet_matches[sheet.title]["nni"].add(match)
                            # LAGパターンのマッチ
                            for match in LAG_PATTERN.findall(value):
                                sheet_matches[sheet.title]["lag"].add(match)
            except Exception as exc:
                print(f"    ! Error reading sheet '{sheet.title}' in {file_path.name}: {exc}", flush=True)
    finally:
        wb.close()
    
    # 結果を整形
    for sheet_name, matches in sheet_matches.items():
        nni_list = "; ".join(sorted(matches["nni"]))
        lag_list = "; ".join(sorted(matches["lag"]))
        rows.append({
            "file": str(file_path),
            "sheet_name": sheet_name,
            "nni_matches": nni_list,
            "lag_matches": lag_list,
        })
    
    return rows


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    target_dir = script_dir / "PTN_Tunnel_PW"
    if not target_dir.exists():
        print(f"Target folder not found: {target_dir}", file=sys.stderr)
        sys.exit(1)

    excel_files = collect_excel_files(target_dir)
    if not excel_files:
        print("No .xlsx or .xlsm files found.")
        return

    print(f"Found {len(excel_files)} Excel files. Starting processing...")
    summary_rows = []
    for idx, file_path in enumerate(excel_files, start=1):
        print(f"[{idx}/{len(excel_files)}] Processing {file_path}", flush=True)
        summary_rows.extend(extract_sheet_data(file_path))

    output_file = script_dir / "ringcheck_summary.csv"
    with output_file.open("w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=["file", "sheet_name", "nni_matches", "lag_matches"])
        writer.writeheader()
        writer.writerows(summary_rows)

    print(f"Done. Summary saved to {output_file}")


if __name__ == "__main__":
    main()