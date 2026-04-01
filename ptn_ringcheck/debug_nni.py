#!/usr/bin/env python3
"""Debug script to check what text is in the Excel file."""

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook

file_path = Path("/Users/Taishi.Kurosaki/Documents/repo01/repo01/ptn_ringcheck/PTN_Tunnel_PW/PTN_Tunnel_PW(中国).xlsx")

print("=== Drawing XML ===")
try:
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        drawing_files = [n for n in zip_ref.namelist() if 'xl/drawings/drawing' in n and n.endswith('.xml')]
        print(f"Found {len(drawing_files)} drawing files: {drawing_files}")
        
        for drawing_file in drawing_files:
            content = zip_ref.read(drawing_file).decode('utf-8', errors='replace')
            root = ET.fromstring(content)
            texts = []
            for elem in root.iter():
                if elem.text and isinstance(elem.text, str):
                    text = elem.text.strip()
                    if text and len(text) > 3:  # 短すぎるテキストは除外
                        texts.append(text)
            print(f"\n{drawing_file}: {len(texts)} texts")
            for i, text in enumerate(texts[:20], 1):  # 最初の20個を表示
                print(f"  {i}. {text[:100]}")  # 最初の100文字のみ
except Exception as exc:
    print(f"Error: {exc}")

print("\n=== Cell Values (first sheet) ===")
wb = load_workbook(filename=file_path, read_only=True, data_only=True)
sheet = wb.worksheets[0]
print(f"Sheet name: {sheet.title}")

cell_values = []
for row in sheet.iter_rows(values_only=True):
    for value in row:
        if isinstance(value, str) and len(value) > 5:
            cell_values.append(value)

print(f"Found {len(cell_values)} string cells")
for i, value in enumerate(cell_values[:20], 1):
    print(f"  {i}. {value[:100]}")

wb.close()

print("\n=== Search for 'NNI' or 'LAG' ===")
# 'NNI'または'LAG'を含むテキストを探す
wb = load_workbook(filename=file_path, read_only=True, data_only=True)
for sheet in wb.worksheets:
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and ('NNI' in value.upper() or 'LAG' in value.upper()):
                print(f"  [{sheet.title}] {value}")
wb.close()
